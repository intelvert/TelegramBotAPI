import telebot
from config import TOKEN 
from openpyxl import load_workbook
import os

bot = telebot.TeleBot(TOKEN)
print(os.getcwd())
excel_file = "eisenhower_matrix.xlsx"

@bot.message_handler(func=lambda message: True)
def handle_message(message):
    title=''
    category = 0
    task = message.text.split('@')[0]
    if '@1' in message.text.lower():
        title = 'Urgent & Important'
        category=1
    elif '@2' in message.text.lower():
        title = 'Not Urgent & Important'
        category = 2
    elif '@3' in message.text.lower():
        title = 'Urgent & Not Important'
        category = 3
    elif '@4' in message.text.lower():
        title = 'Not Urgent & Not Important'
        category = 4

    # saving task into Excel
    if category:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        # Finding first empty cell
        for row in range(1, sheet.max_row + 2):
            if sheet.cell(row=row, column=category).value is None:
                sheet.cell(row=row, column=category, value=task)
                break
        workbook.save(excel_file)

        bot.send_message(chat_id=message.chat.id, reply_to_message_id=message.message_id, text="📝")
    else:
        bot.reply_to(message, "Unable to detect task category")

bot.polling(none_stop=True)


